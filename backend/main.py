from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from typing import Optional, List
import subprocess
import json
import sqlite3
from datetime import datetime
import os
import asyncio
import time
from io import BytesIO
from urllib import error as urllib_error
from urllib import request as urllib_request

from dotenv import load_dotenv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

app = FastAPI(title="CI/CD Security Dashboard API")

# CORS middleware for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database initialization
configured_db_path = os.getenv("DATABASE_PATH", "vulnerabilities.db")
DATABASE_PATH = configured_db_path if os.path.isabs(configured_db_path) else os.path.join(BASE_DIR, configured_db_path)
TRIVY_TIMEOUT_SECONDS = int(os.getenv("TRIVY_TIMEOUT", "300"))
TRIVY_SKIP_DB_UPDATE = os.getenv("TRIVY_SKIP_DB_UPDATE", "true").strip().lower() in {"1", "true", "yes"}
TRIVY_CACHE_DIR = (os.getenv("TRIVY_CACHE_DIR") or "").strip()
DOCKER_HOST = (os.getenv("DOCKER_HOST") or "unix:///var/run/docker.sock").strip()


@app.on_event("startup")
async def log_startup_context():
    print(f"[startup] Using SQLite database: {DATABASE_PATH}")
    print(f"[startup] Docker host: {DOCKER_HOST}")
    slack_webhook = (os.getenv("SLACK_WEBHOOK_URL") or "").strip()
    slack_bot_token = (os.getenv("SLACK_BOT_TOKEN") or "").strip()
    slack_channel_id = (os.getenv("SLACK_CHANNEL_ID") or "").strip()
    webhook_configured = bool(slack_webhook)
    bot_configured = bool(slack_bot_token and slack_channel_id)
    print(f"[startup] Slack webhook configured: {webhook_configured}")
    print(f"[startup] Slack bot fallback configured: {bot_configured}")


def init_db():
    """Initialize SQLite database with required tables"""
    conn = sqlite3.connect(DATABASE_PATH, timeout=10)
    cursor = conn.cursor()
    cursor.execute("PRAGMA busy_timeout = 10000")
    cursor.execute("PRAGMA journal_mode = WAL")
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS scans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            image_name TEXT NOT NULL,
            scan_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            scan_source TEXT DEFAULT 'remote',
            total_vulnerabilities INTEGER DEFAULT 0,
            critical_count INTEGER DEFAULT 0,
            high_count INTEGER DEFAULT 0,
            medium_count INTEGER DEFAULT 0,
            low_count INTEGER DEFAULT 0
        )
    """)

    # Backward-compatible schema migration for older DB files.
    cursor.execute("PRAGMA table_info(scans)")
    scan_columns = [row[1] for row in cursor.fetchall()]
    if "scan_source" not in scan_columns:
        cursor.execute("ALTER TABLE scans ADD COLUMN scan_source TEXT DEFAULT 'remote'")
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS vulnerabilities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scan_id INTEGER NOT NULL,
            package_name TEXT,
            installed_version TEXT,
            fixed_version TEXT,
            vulnerability_id TEXT,
            severity TEXT,
            description TEXT,
            FOREIGN KEY (scan_id) REFERENCES scans(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    conn.commit()
    conn.close()


# Initialize database on startup
init_db()


class ScanRequest(BaseModel):
    image_name: str
    alert_email: Optional[str] = None


class VulnerabilityResponse(BaseModel):
    id: int
    scan_id: int
    package_name: Optional[str]
    installed_version: Optional[str]
    fixed_version: Optional[str]
    vulnerability_id: Optional[str]
    severity: str
    description: Optional[str]


class ScanSummary(BaseModel):
    id: int
    image_name: str
    scan_date: str
    total_vulnerabilities: int
    critical_count: int
    high_count: int
    medium_count: int
    low_count: int


def normalize_email(email: str):
    return (email or "").strip().lower()


def is_valid_email(email: str):
    normalized = normalize_email(email)
    return "@" in normalized and "." in normalized.split("@")[-1]


def resolve_alert_email(request_email: Optional[str]):
    if request_email:
        normalized = normalize_email(request_email)
        return normalized if is_valid_email(normalized) else None

    return None


def send_vulnerability_alert_slack(
    recipient_label: str,
    image_name: str,
    scan_id: int,
    severity_counts: dict,
    scan_source: str,
):
    slack_webhook_url = (os.getenv("SLACK_WEBHOOK_URL") or "").strip()
    slack_bot_token = (os.getenv("SLACK_BOT_TOKEN") or "").strip()
    slack_channel_id = (os.getenv("SLACK_CHANNEL_ID") or "").strip()
    slack_alert_label = (os.getenv("SLACK_ALERT_LABEL") or "#security-alerts").strip()

    target_label = slack_alert_label if recipient_label == "slack" else recipient_label
    source_label = "Local Image Scanner" if scan_source == "local" else "Pre-built Image Scanner"
    message_text = (
        "*VULNERABILITY DETECTED*\n"
        f"*Source:* {source_label}\n"
        f"*Image:* {image_name}\n"
        f"*Scan ID:* {scan_id}\n"
        f"*Target:* {target_label}\n"
        f"*Critical:* {severity_counts.get('CRITICAL', 0)}\n"
        f"*High:* {severity_counts.get('HIGH', 0)}\n"
        f"*Medium:* {severity_counts.get('MEDIUM', 0)}\n"
        f"*Low:* {severity_counts.get('LOW', 0)}\n"
        f"*Timestamp:* {datetime.utcnow().isoformat()}Z"
    )

    webhook_error = None

    # Primary path: Incoming webhook.
    if slack_webhook_url and not slack_webhook_url.startswith("your-"):
        try:
            webhook_req = urllib_request.Request(
                slack_webhook_url,
                data=json.dumps({"text": message_text}).encode("utf-8"),
                headers={"Content-Type": "application/json"},
                method="POST",
            )

            with urllib_request.urlopen(webhook_req, timeout=15) as response:
                response_body = response.read().decode("utf-8", errors="replace").strip()

            if response_body in {"ok", ""}:
                return True, None

            webhook_error = f"Slack webhook response: {response_body}"
        except urllib_error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace") if exc.fp else str(exc)
            webhook_error = f"Slack webhook error: HTTP {exc.code} {detail}"
        except Exception as exc:
            webhook_error = f"Slack webhook error: {str(exc)}"

    # Fallback path: Bot token + chat.postMessage.
    if slack_bot_token and slack_channel_id and not slack_bot_token.startswith("xoxb-your-"):
        try:
            bot_payload = {
                "channel": slack_channel_id,
                "text": message_text,
            }
            bot_req = urllib_request.Request(
                "https://slack.com/api/chat.postMessage",
                data=json.dumps(bot_payload).encode("utf-8"),
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {slack_bot_token}",
                },
                method="POST",
            )

            with urllib_request.urlopen(bot_req, timeout=15) as response:
                raw_body = response.read().decode("utf-8", errors="replace")

            parsed = json.loads(raw_body or "{}")
            if parsed.get("ok"):
                return True, None

            bot_error = f"Slack bot error: {parsed.get('error', 'unknown_error')}"
            if webhook_error:
                return False, f"{webhook_error} | {bot_error}"
            return False, bot_error
        except urllib_error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace") if exc.fp else str(exc)
            bot_http_error = f"Slack bot HTTP error: HTTP {exc.code} {detail}"
            if webhook_error:
                return False, f"{webhook_error} | {bot_http_error}"
            return False, bot_http_error
        except Exception as exc:
            bot_error = f"Slack bot error: {str(exc)}"
            if webhook_error:
                return False, f"{webhook_error} | {bot_error}"
            return False, bot_error

    if webhook_error:
        return False, webhook_error

    return False, "Slack is not configured. Set SLACK_WEBHOOK_URL or SLACK_BOT_TOKEN + SLACK_CHANNEL_ID in backend/.env."


def build_trivy_scan_command(image_reference: str, skip_db_update: bool):
    cmd = [
        "trivy", "image",
        "--scanners", "vuln",
        "--format", "json",
        "--quiet",
    ]

    if TRIVY_CACHE_DIR:
        cmd.extend(["--cache-dir", TRIVY_CACHE_DIR])

    if skip_db_update:
        # Fast mode for repeat scans; fallback retry without these flags happens on failure.
        cmd.extend(["--skip-db-update", "--skip-java-db-update"])

    cmd.append(image_reference)
    return cmd


def get_subprocess_env():
    env = os.environ.copy()
    env["DOCKER_HOST"] = DOCKER_HOST
    return env


def run_trivy_scan(image_reference: str):
    initial_cmd = build_trivy_scan_command(image_reference, TRIVY_SKIP_DB_UPDATE)
    result = subprocess.run(
        initial_cmd,
        capture_output=True,
        text=True,
        timeout=TRIVY_TIMEOUT_SECONDS,
        env=get_subprocess_env(),
    )

    if result.returncode == 0 or not TRIVY_SKIP_DB_UPDATE:
        return result

    stderr_output = (result.stderr or "").lower()
    first_run_db_error = "cannot be specified on the first run" in stderr_output
    db_metadata_error = "database error" in stderr_output

    # Retry once without skip flags when cache metadata is not initialized yet.
    if not (first_run_db_error or db_metadata_error):
        return result

    fallback_cmd = build_trivy_scan_command(image_reference, skip_db_update=False)
    return subprocess.run(
        fallback_cmd,
        capture_output=True,
        text=True,
        timeout=TRIVY_TIMEOUT_SECONDS,
        env=get_subprocess_env(),
    )


def get_local_docker_images():
    """Return local Docker images from the host daemon."""
    result = subprocess.run(
        ["docker", "image", "ls", "--format", "{{json .}}"],
        capture_output=True,
        text=True,
        timeout=30,
        env=get_subprocess_env(),
    )

    if result.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to list local Docker images: {result.stderr.strip()}"
        )

    images = []
    for line in result.stdout.splitlines():
        line = line.strip()
        if not line:
            continue

        parsed = json.loads(line)
        repository = parsed.get("Repository", "<none>")
        tag = parsed.get("Tag", "<none>")
        image_name = f"{repository}:{tag}" if tag != "<none>" else repository

        images.append({
            "image_name": image_name,
            "repository": repository,
            "tag": tag,
            "image_id": parsed.get("ID", ""),
            "created_since": parsed.get("CreatedSince", ""),
            "size": parsed.get("Size", "")
        })

    return images


def resolve_local_image_reference(image_name: str):
    """Resolve a local image name to a usable scan reference (name or image ID)."""
    inspect_result = subprocess.run(
        ["docker", "image", "inspect", image_name],
        capture_output=True,
        text=True,
        timeout=30,
        env=get_subprocess_env(),
    )

    if inspect_result.returncode == 0:
        return image_name, ""

    # Fallback: resolve by image list in case tag inspect is inconsistent.
    list_result = subprocess.run(
        ["docker", "image", "ls", "--no-trunc", "--format", "{{.Repository}}:{{.Tag}} {{.ID}}"],
        capture_output=True,
        text=True,
        timeout=30,
        env=get_subprocess_env(),
    )

    if list_result.returncode == 0:
        for line in list_result.stdout.splitlines():
            parts = line.split()
            if len(parts) != 2:
                continue

            listed_name, listed_id = parts
            if listed_name == image_name:
                return listed_id, ""

    return None, inspect_result.stderr.strip()


def fetch_scan_with_vulnerabilities(scan_id: int):
    """Load one scan row and all vulnerabilities for report generation."""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM scans WHERE id = ?", (scan_id,))
    scan_row = cursor.fetchone()

    if not scan_row:
        conn.close()
        return None, []

    cursor.execute("""
        SELECT * FROM vulnerabilities
        WHERE scan_id = ?
        ORDER BY id ASC
    """, (scan_id,))
    vulnerabilities = [dict(row) for row in cursor.fetchall()]

    conn.close()
    return dict(scan_row), vulnerabilities


def build_pdf_report(scan: dict, vulnerabilities: List[dict]):
    """Generate a PDF scan report as bytes."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    source_label = "Local Image Scanner" if scan.get("scan_source") == "local" else "Docker Image Scanner"
    elements.append(Paragraph("Security Scan Report", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Source: {source_label}", styles["Normal"]))
    elements.append(Paragraph(f"Image: {scan.get('image_name')}", styles["Normal"]))
    elements.append(Paragraph(f"Scan ID: {scan.get('id')}", styles["Normal"]))
    elements.append(Paragraph(f"Scan Date: {scan.get('scan_date')}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    summary_table = Table([
        ["Critical", scan.get("critical_count", 0)],
        ["High", scan.get("high_count", 0)],
        ["Medium", scan.get("medium_count", 0)],
        ["Low", scan.get("low_count", 0)],
        ["Total", scan.get("total_vulnerabilities", 0)],
    ], colWidths=[150, 120])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Vulnerabilities", styles["Heading2"]))
    if not vulnerabilities:
        elements.append(Paragraph("No vulnerabilities found.", styles["Normal"]))
    else:
        table_cell_style = styles["BodyText"].clone("vulnTableCell")
        table_cell_style.fontName = "Helvetica"
        table_cell_style.fontSize = 8
        table_cell_style.leading = 10
        table_cell_style.wordWrap = "CJK"

        rows = [["ID", "Package", "Severity", "Installed", "Fixed"]]
        for vuln in vulnerabilities:
            rows.append([
                Paragraph(str(vuln.get("vulnerability_id") or "-"), table_cell_style),
                Paragraph(str(vuln.get("package_name") or "-"), table_cell_style),
                Paragraph(str(vuln.get("severity") or "-"), table_cell_style),
                Paragraph(str(vuln.get("installed_version") or "-"), table_cell_style),
                Paragraph(str(vuln.get("fixed_version") or "-"), table_cell_style),
            ])

        vuln_table = Table(
            rows,
            repeatRows=1,
            colWidths=[120, 80, 70, 90, 90],
        )
        vuln_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(vuln_table)

        elements.append(Spacer(1, 14))
        elements.append(Paragraph("Vulnerability Descriptions", styles["Heading2"]))
        details_style = styles["BodyText"].clone("detailsText")
        details_style.fontName = "Helvetica"
        details_style.fontSize = 9
        details_style.leading = 12
        details_style.wordWrap = "CJK"

        for vuln in vulnerabilities:
            vuln_id = str(vuln.get("vulnerability_id") or "-")
            pkg = str(vuln.get("package_name") or "-")
            desc = str(vuln.get("description") or "-")

            title = f"<b>{vuln_id}</b> ({pkg})"
            elements.append(Paragraph(title, details_style))
            elements.append(Paragraph(desc, details_style))
            elements.append(Spacer(1, 8))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def build_excel_report(scan: dict, vulnerabilities: List[dict]):
    """Generate an Excel scan report as bytes."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    source_label = "Local Image Scanner" if scan.get("scan_source") == "local" else "Docker Image Scanner"

    ws_summary.append(["Security Scan Report"])
    ws_summary.append([])
    ws_summary.append(["Source", source_label])
    ws_summary.append(["Image", scan.get("image_name")])
    ws_summary.append(["Scan ID", scan.get("id")])
    ws_summary.append(["Scan Date", scan.get("scan_date")])
    ws_summary.append([])
    ws_summary.append(["Critical", scan.get("critical_count", 0)])
    ws_summary.append(["High", scan.get("high_count", 0)])
    ws_summary.append(["Medium", scan.get("medium_count", 0)])
    ws_summary.append(["Low", scan.get("low_count", 0)])
    ws_summary.append(["Total", scan.get("total_vulnerabilities", 0)])

    ws_summary.append([])
    ws_summary.append(["Vulnerability ID", "Package", "Severity", "Installed Version", "Fixed Version", "Description"])
    if vulnerabilities:
        for vuln in vulnerabilities:
            ws_summary.append([
                vuln.get("vulnerability_id") or "-",
                vuln.get("package_name") or "-",
                vuln.get("severity") or "-",
                vuln.get("installed_version") or "-",
                vuln.get("fixed_version") or "-",
                vuln.get("description") or "-",
            ])
    else:
        ws_summary.append(["-", "-", "-", "-", "-", "No vulnerabilities found."])

    ws_vuln = wb.create_sheet(title="Vulnerabilities")
    ws_vuln.append(["Vulnerability ID", "Package", "Severity", "Installed Version", "Fixed Version", "Description"])
    if vulnerabilities:
        for vuln in vulnerabilities:
            ws_vuln.append([
                vuln.get("vulnerability_id") or "-",
                vuln.get("package_name") or "-",
                vuln.get("severity") or "-",
                vuln.get("installed_version") or "-",
                vuln.get("fixed_version") or "-",
                vuln.get("description") or "-",
            ])
    else:
        ws_vuln.append(["-", "-", "-", "-", "-", "No vulnerabilities found."])

    # Format sheets for readability and long description wrapping.
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)

    for ws in (ws_summary, ws_vuln):
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 80

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary_header_row = 14
    for col in range(1, 7):
        cell = ws_summary.cell(row=summary_header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for col in range(1, 7):
        cell = ws_vuln.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    ws_summary.freeze_panes = "A15"
    ws_vuln.freeze_panes = "A2"

    stream = BytesIO()
    wb.save(stream)
    excel_bytes = stream.getvalue()
    stream.close()
    return excel_bytes


@app.get("/")
def read_root():
    return {"message": "CI/CD Security Dashboard API", "version": "1.0"}


@app.post("/scan")
async def scan_image(request: ScanRequest):
    """
    Scan a Docker image using Trivy and store results in SQLite
    """
    image_name = request.image_name
    scan_started = time.monotonic()
    
    try:
        # Skip remote pull if image already exists locally.
        local_scan_reference, _ = resolve_local_image_reference(image_name)
        scan_reference = local_scan_reference or image_name

        if not local_scan_reference:
            print(f"Pulling image: {image_name}")
            pull_result = subprocess.run(
                ["docker", "pull", image_name],
                capture_output=True,
                text=True,
                timeout=TRIVY_TIMEOUT_SECONDS,
                env=get_subprocess_env(),
            )

            if pull_result.returncode != 0:
                print(f"Warning: Failed to pull image: {pull_result.stderr}")
                # Continue anyway - image might already exist locally

        # Run Trivy scan
        print(f"Scanning image: {image_name}")
        result = run_trivy_scan(scan_reference)
        
        if result.returncode != 0:
            raise HTTPException(
                status_code=500,
                detail=f"Trivy scan failed: {result.stderr}"
            )
        
        # Parse JSON results
        scan_results = json.loads(result.stdout)
        
        # Extract vulnerabilities
        vulnerabilities = []
        for item in scan_results.get("Results", []):
            vulns = item.get("Vulnerabilities", [])
            if vulns:
                vulnerabilities.extend(vulns)
        
        # Count by severity
        severity_counts = {
            "CRITICAL": 0,
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0
        }
        
        for vuln in vulnerabilities:
            severity = vuln.get("Severity", "UNKNOWN")
            if severity in severity_counts:
                severity_counts[severity] += 1
        
        # Store in database
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        
        # Insert scan record
        cursor.execute("""
            INSERT INTO scans (
                image_name, scan_source, total_vulnerabilities,
                critical_count, high_count, medium_count, low_count
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            image_name,
            "remote",
            len(vulnerabilities),
            severity_counts["CRITICAL"],
            severity_counts["HIGH"],
            severity_counts["MEDIUM"],
            severity_counts["LOW"]
        ))
        
        scan_id = cursor.lastrowid
        
        # Insert vulnerabilities in one batch for better write performance.
        vuln_rows = [
            (
                scan_id,
                vuln.get("PkgName"),
                vuln.get("InstalledVersion"),
                vuln.get("FixedVersion"),
                vuln.get("VulnerabilityID"),
                vuln.get("Severity"),
                vuln.get("Description"),
            )
            for vuln in vulnerabilities
        ]

        if vuln_rows:
            cursor.executemany("""
                INSERT INTO vulnerabilities (
                    scan_id, package_name, installed_version,
                    fixed_version, vulnerability_id, severity, description
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, vuln_rows)
        
        conn.commit()
        conn.close()

        should_alert = severity_counts["CRITICAL"] > 0 or severity_counts["HIGH"] > 0
        alert_sent = False
        alert_error = None
        alert_email = resolve_alert_email(request.alert_email)

        if should_alert:
            alert_sent, alert_error = send_vulnerability_alert_slack(
                recipient_label="slack",
                image_name=image_name,
                scan_id=scan_id,
                severity_counts=severity_counts,
                scan_source="remote",
            )
        
        return {
            "success": True,
            "scan_id": scan_id,
            "image_name": image_name,
            "total_vulnerabilities": len(vulnerabilities),
            "severity_counts": severity_counts,
            "message": "Scan completed successfully",
            "alert_sent": alert_sent,
            "alert_error": alert_error,
            "alert_target": "slack" if should_alert else None,
            "scan_duration_seconds": round(time.monotonic() - scan_started, 2)
        }
        
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=500, detail="Scan timed out")
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Failed to parse Trivy output")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.post("/scan/local")
async def scan_local_image(request: ScanRequest):
    """
    Scan a local Docker image using Trivy and store results in SQLite.
    This endpoint never pulls from a remote registry.
    """
    image_name = request.image_name
    scan_started = time.monotonic()

    try:
        # Ensure the requested image exists locally before scanning.
        scan_reference, inspect_error = resolve_local_image_reference(image_name)

        if not scan_reference:
            raise HTTPException(
                status_code=404,
                detail=f"Local image not found: {image_name}. Docker said: {inspect_error}"
            )

        # Run Trivy scan directly against local image.
        result = run_trivy_scan(scan_reference)

        if result.returncode != 0:
            raise HTTPException(
                status_code=500,
                detail=f"Trivy scan failed: {result.stderr}"
            )

        scan_results = json.loads(result.stdout)

        vulnerabilities = []
        for item in scan_results.get("Results", []):
            vulns = item.get("Vulnerabilities", [])
            if vulns:
                vulnerabilities.extend(vulns)

        severity_counts = {
            "CRITICAL": 0,
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0
        }

        for vuln in vulnerabilities:
            severity = vuln.get("Severity", "UNKNOWN")
            if severity in severity_counts:
                severity_counts[severity] += 1

        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO scans (
                image_name, scan_source, total_vulnerabilities,
                critical_count, high_count, medium_count, low_count
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            image_name,
            "local",
            len(vulnerabilities),
            severity_counts["CRITICAL"],
            severity_counts["HIGH"],
            severity_counts["MEDIUM"],
            severity_counts["LOW"]
        ))

        scan_id = cursor.lastrowid

        vuln_rows = [
            (
                scan_id,
                vuln.get("PkgName"),
                vuln.get("InstalledVersion"),
                vuln.get("FixedVersion"),
                vuln.get("VulnerabilityID"),
                vuln.get("Severity"),
                vuln.get("Description"),
            )
            for vuln in vulnerabilities
        ]

        if vuln_rows:
            cursor.executemany("""
                INSERT INTO vulnerabilities (
                    scan_id, package_name, installed_version,
                    fixed_version, vulnerability_id, severity, description
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, vuln_rows)

        conn.commit()
        conn.close()

        should_alert = severity_counts["CRITICAL"] > 0 or severity_counts["HIGH"] > 0
        alert_sent = False
        alert_error = None
        alert_email = resolve_alert_email(request.alert_email)

        if should_alert:
            alert_sent, alert_error = send_vulnerability_alert_slack(
                recipient_label="slack",
                image_name=image_name,
                scan_id=scan_id,
                severity_counts=severity_counts,
                scan_source="local",
            )

        return {
            "success": True,
            "scan_id": scan_id,
            "image_name": image_name,
            "total_vulnerabilities": len(vulnerabilities),
            "severity_counts": severity_counts,
            "message": "Local image scan completed successfully",
            "alert_sent": alert_sent,
            "alert_error": alert_error,
            "alert_target": "slack" if should_alert else None,
            "scan_duration_seconds": round(time.monotonic() - scan_started, 2)
        }

    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=500, detail="Scan timed out")
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Failed to parse Trivy output")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/reports/scan/{scan_id}")
async def download_scan_report(scan_id: int, format: str = "pdf"):
    """Download a single scan report as PDF or Excel."""
    fmt = format.lower()
    if fmt not in {"pdf", "excel"}:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'pdf' or 'excel'.")

    scan, vulnerabilities = fetch_scan_with_vulnerabilities(scan_id)
    if not scan:
        raise HTTPException(status_code=404, detail=f"Scan not found: {scan_id}")

    source = scan.get("scan_source") or "remote"
    base_name = f"scan-report-{source}-{scan_id}"

    if fmt == "pdf":
        content = build_pdf_report(scan, vulnerabilities)
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={base_name}.pdf"},
        )

    content = build_excel_report(scan, vulnerabilities)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={base_name}.xlsx"},
    )


@app.get("/docker/images")
async def list_local_images():
    """List local Docker images available on the host machine."""
    try:
        images = get_local_docker_images()
        return {
            "success": True,
            "count": len(images),
            "images": images
        }
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Failed to parse Docker CLI output")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/docker/events")
async def stream_docker_image_events():
    """Stream Docker image-related daemon events for real-time frontend updates."""

    async def event_generator():
        process = await asyncio.create_subprocess_exec(
            "docker", "events",
            "--filter", "type=image",
            "--format", "{{json .}}",
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )

        try:
            # Initial event lets the UI mark the connection as ready.
            yield f"data: {json.dumps({'type': 'connected'})}\n\n"

            while True:
                line = await process.stdout.readline()
                if not line:
                    break

                payload = line.decode("utf-8").strip()
                if not payload:
                    continue

                try:
                    event_data = json.loads(payload)
                    event_name = event_data.get("Action") or event_data.get("status") or "unknown"
                    event_time = event_data.get("timeNano") or event_data.get("time")

                    outgoing = {
                        "type": "image_event",
                        "action": event_name,
                        "time": event_time,
                    }
                except json.JSONDecodeError:
                    outgoing = {
                        "type": "raw_event",
                        "payload": payload,
                    }

                yield f"data: {json.dumps(outgoing)}\n\n"
        finally:
            if process.returncode is None:
                process.terminate()
                await process.wait()

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/results")
async def get_results(severity: Optional[str] = None, limit: int = 100):
    """
    Get vulnerability results with optional severity filtering
    """
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        if severity:
            severity = severity.upper()
            cursor.execute("""
                SELECT * FROM vulnerabilities 
                WHERE severity = ? 
                ORDER BY id DESC 
                LIMIT ?
            """, (severity, limit))
        else:
            cursor.execute("""
                SELECT * FROM vulnerabilities 
                ORDER BY id DESC 
                LIMIT ?
            """, (limit,))
        
        vulnerabilities = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return {
            "success": True,
            "count": len(vulnerabilities),
            "vulnerabilities": vulnerabilities
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/scans")
async def get_scans(limit: int = 20, scan_id: Optional[int] = None):
    """
    Get scan history, optionally filtered by scan ID.
    """
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        if scan_id is not None:
            cursor.execute("""
                SELECT * FROM scans
                WHERE id = ?
                ORDER BY scan_date DESC
            """, (scan_id,))
        else:
            cursor.execute("""
                SELECT * FROM scans
                ORDER BY scan_date DESC
                LIMIT ?
            """, (limit,))
        
        scans = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return {
            "success": True,
            "count": len(scans),
            "scans": scans
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/scans/{scan_id}")
async def get_scan_details(scan_id: int):
    """
    Get one scan and all associated vulnerabilities.
    """
    try:
        scan, vulnerabilities = fetch_scan_with_vulnerabilities(scan_id)
        if not scan:
            raise HTTPException(status_code=404, detail=f"Scan not found: {scan_id}")

        return {
            "success": True,
            "scan": scan,
            "count": len(vulnerabilities),
            "vulnerabilities": vulnerabilities,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/stats")
async def get_stats():
    """
    Get overall statistics for dashboard
    """
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get total scans
        cursor.execute("SELECT COUNT(*) as count FROM scans")
        total_scans = cursor.fetchone()["count"]
        
        # Get total vulnerabilities
        cursor.execute("SELECT COUNT(*) as count FROM vulnerabilities")
        total_vulnerabilities = cursor.fetchone()["count"]
        
        # Get severity distribution across all scans
        cursor.execute("""
            SELECT severity, COUNT(*) as count 
            FROM vulnerabilities 
            GROUP BY severity
        """)
        severity_distribution = {row["severity"]: row["count"] for row in cursor.fetchall()}

        # Get severity distribution for local image scans
        cursor.execute("""
            SELECT v.severity, COUNT(*) as count
            FROM vulnerabilities v
            JOIN scans s ON s.id = v.scan_id
            WHERE s.scan_source = 'local'
            GROUP BY v.severity
        """)
        local_severity_distribution = {row["severity"]: row["count"] for row in cursor.fetchall()}

        # Get severity distribution for remote Docker image scans
        cursor.execute("""
            SELECT v.severity, COUNT(*) as count
            FROM vulnerabilities v
            JOIN scans s ON s.id = v.scan_id
            WHERE s.scan_source IS NULL OR s.scan_source != 'local'
            GROUP BY v.severity
        """)
        remote_severity_distribution = {row["severity"]: row["count"] for row in cursor.fetchall()}
        
        # Get recent scans for timeline
        cursor.execute("""
            SELECT scan_date, total_vulnerabilities, critical_count, 
                   high_count, medium_count, low_count, image_name,
                   COALESCE(scan_source, 'remote') as scan_source
            FROM scans 
            ORDER BY scan_date DESC 
            LIMIT 10
        """)
        recent_scans = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return {
            "success": True,
            "total_scans": total_scans,
            "total_vulnerabilities": total_vulnerabilities,
            "severity_distribution": severity_distribution,
            "local_severity_distribution": local_severity_distribution,
            "remote_severity_distribution": remote_severity_distribution,
            "recent_scans": recent_scans
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
